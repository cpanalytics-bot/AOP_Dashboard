"""Builds one consolidated, plain-English Excel workbook from the discovery CSVs."""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "AOP_Data_Architecture.xlsx")

# ---- styling -------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")   # dark slate
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=15, color="111827")
SUB_FONT = Font(italic=True, size=10, color="6B7280")
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "CONFIRMED": PatternFill("solid", fgColor="D1FAE5"),     # green
    "DERIVED": PatternFill("solid", fgColor="DBEAFE"),       # blue
    "LOOKUP": PatternFill("solid", fgColor="CCFBF1"),        # teal
    "NEEDS_MAPPING": PatternFill("solid", fgColor="FECACA"), # red
    "NEEDS_INPUT": PatternFill("solid", fgColor="FEF3C7"),   # amber
    "N/A": PatternFill("solid", fgColor="F3F4F6"),
}

wb = Workbook()


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, max_w=60, start_row=1):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        longest = 0
        for row in range(start_row, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            longest = max(longest, max(len(line) for line in str(v).split("\n")))
        ws.column_dimensions[letter].width = min(max(12, longest + 2), max_w)


def add_intro(ws, title, subtitle, header_row):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = SUB_FONT
    ws.cell(row=2, column=1).alignment = WRAP
    return header_row


def color_status(ws, status_col_idx, first_data_row):
    for row in range(first_data_row, ws.max_row + 1):
        val = ws.cell(row=row, column=status_col_idx).value
        if val and val in STATUS_FILL:
            ws.cell(row=row, column=status_col_idx).fill = STATUS_FILL[val]


def write_csv_sheet(sheet_name, csv_name, title, subtitle, status_col=None):
    ws = wb.create_sheet(sheet_name)
    add_intro(ws, title, subtitle, 4)
    with open(os.path.join(HERE, csv_name), newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    header = rows[0]
    hr = 4
    for c, h in enumerate(header, 1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, len(header), row=hr)
    for r, row in enumerate(rows[1:], hr + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
    if status_col is not None:
        idx = header.index(status_col) + 1
        color_status(ws, idx, hr + 1)
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    autosize(ws, start_row=hr)
    return ws


def write_plain_sheet(sheet_name, title, subtitle, columns, data):
    ws = wb.create_sheet(sheet_name)
    add_intro(ws, title, subtitle, 4)
    hr = 4
    for c, h in enumerate(columns, 1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, len(columns), row=hr)
    for r, row in enumerate(data, hr + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    autosize(ws, start_row=hr)
    return ws


# ==========================================================================
# Sheet 1: Start Here (plain english overview)
# ==========================================================================
ws = wb.active
ws.title = "Start Here"
ws.cell(row=1, column=1, value="AOP Platform - Data Map (Plain English)").font = Font(bold=True, size=18, color="111827")
intro = [
    ("What is this file?",
     "This workbook lists EVERY value shown anywhere in the AOP app, and explains in plain words where each value should come from. It is a checklist to agree on data BEFORE we build the database."),
    ("Why does it matter?",
     "Right now the app shows some numbers that are FAKE (placeholders we made up for the demo). Before go-live we must replace them with real data from a real system. This file flags exactly which ones."),
    ("How to read the colours (the 'Status' column)",
     ""),
    ("   CONFIRMED (green)",
     "Good. The user types this in, and we already have a clear place to store it."),
    ("   DERIVED (blue)",
     "The app calculates this from other values. We don't store it; we compute it. The formula is known."),
    ("   LOOKUP (teal)",
     "We fetch this by matching to a master list (like the employee or territory list)."),
    ("   NEEDS_MAPPING (red)",
     "DANGER. This value is currently fake / made-up. We do NOT yet know the real source. Business must tell us."),
    ("   NEEDS_INPUT (amber)",
     "The design exists, but a rule, number, or owner still needs to be confirmed by the business."),
    ("What are the tabs?",
     "Screens = every page. All Fields = every box on every page. Where Values Come From = the source for each. KPIs / Calculations = every number we compute. Historical Data = past numbers we need. Pre-fill = which boxes auto-fill. Existing/New Tables = the database. Relationships = how tables link. Big Risks = the fake values. Open Questions = what we need YOU to answer."),
    ("The 6 biggest risks (read the 'Big Risks' tab)",
     "1) Last year's revenue is guessed from a ratio. 2) Average order value is a fixed 1.45 lakh for everyone. 3) Revenue per school is a fixed 2.4 lakh for everyone. 4) School counts are made-up numbers. 5) The dashboard 'achievement %' is a random demo formula. 6) Collection % per region is hard-coded."),
    ("What we need from you",
     "Go to the 'Open Questions' tab. There are 18 questions. Each one unlocks one or more red (NEEDS_MAPPING) values. Answer those and we can start building safely."),
]
r = 3
for k, v in intro:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=11, color="111827")
    ws.cell(row=r, column=1).alignment = TOP
    ws.cell(row=r, column=2, value=v).alignment = WRAP
    r += 1
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 95
for rr in range(3, r):
    ws.row_dimensions[rr].height = 42

# ==========================================================================
# Sheet 2: Screens (Phase 1) - plain english
# ==========================================================================
screens = [
    ["Login", "Sign in / pick who you are", "Everyone", "Pick your name (demo) or email + password", "You are signed in"],
    ["Home (employee cards)", "See a list of people and open their yearly plan", "Everyone (you only see yourself + your team)", "Search box, role filter", "Cards showing name, target hit %, plan status"],
    ["Yearly Plan Wizard", "Build the full year plan in 8 simple steps", "You (your own) or your manager", "All the planning numbers", "A saved plan, live scores, and approvals"],
    ["  Step 1 - Hiring", "Ask for new people before planning", "Managers raise; juniors view", "Role, how many, why, when", "A hiring request"],
    ["  Step 2 - Revenue", "Set this year's sales targets", "Plan owner", "Total target + split by product + average order value", "Growth scores"],
    ["  Step 3 - Schools", "Plan how many schools you will cover", "Plan owner", "School counts, category mix, distributor", "School growth score"],
    ["  Step 4 - Sampling", "Plan free trials and how many will convert", "Plan owner", "Samples per stream + conversion %", "Cost per win"],
    ["  Step 5 - Training", "Plan teacher/principal sessions", "Plan owner", "Number of each session type", "Cost per session"],
    ["  Step 6 - Cost", "List all money you will spend", "Plan owner", "12 cost lines", "Total cost, cost %, ROI"],
    ["  Step 7 - Collection", "How much cash you will collect by when (auto)", "Plan owner (read-only)", "Nothing - auto from target x region %", "Cash milestones (Dec/Mar/Apr/Jun)"],
    ["  Step 8 - Review & Submit", "Check everything and send for approval", "Owner + approver", "Submit / Approve / Reject / Ask changes + comment", "Plan status changes"],
    ["Hiring page", "Manage all hiring requests", "Managers (juniors see own)", "The hiring form; status change (top boss)", "Hiring list + status"],
    ["Dashboard", "See how you / your team are doing vs the plan", "You (self) or manager (team)", "Pick a person", "Scores, target-vs-actual, team table"],
]
write_plain_sheet(
    "1. Screens",
    "Phase 1 - Every screen in the app",
    "A plain-English tour of every page, who uses it, what goes in, and what comes out.",
    ["Screen", "What it is for (simple)", "Who uses it", "What you put in", "What you get out"],
    screens,
)

# ==========================================================================
# CSV-backed sheets
# ==========================================================================
write_csv_sheet("2. All Fields", "field_inventory.csv",
                "Phase 2 - Every field (box) on every screen",
                "Every input box, dropdown, score and table column. 'Editable' = can the user change it. 'Mandatory' = must it be filled.")

write_csv_sheet("3. Where Values Come From", "source_mapping.csv",
                "Phase 3 - The source for every value",
                "For each value: how it is produced today, and which table/column should feed it. RED rows are fake values needing a real source.",
                status_col="Status")

write_csv_sheet("4. KPIs", "kpi_inventory.csv",
                "Phase 4 - Every score (KPI) and its formula",
                "Every headline number, what it means, the exact formula, and whether its inputs are real.",
                status_col="Status")

write_csv_sheet("5. Calculations", "calculation_inventory.csv",
                "Phase 5 - Every calculated value",
                "Every behind-the-scenes calculation, its formula, and any missing inputs.",
                status_col="Status")

write_csv_sheet("6. Historical Data Needed", "historical_data_requirements.csv",
                "Phase 6 - Past data the app needs",
                "Anywhere the app needs last year / year-to-date numbers. Most have NO confirmed source yet.",
                status_col="Status")

write_csv_sheet("7. Pre-fill Logic", "prefill_logic_inventory.csv",
                "Phase 7 - Which boxes fill themselves",
                "For each field: is it typed by hand, auto-filled, calculated, or looked up - and from where.",
                status_col="Status")

write_csv_sheet("8. Existing Tables", "existing_tables.csv",
                "Phase 8 - Database tables already designed",
                "The 18 tables (plus a view) already designed in the schema, what they store, and which screens use them.")

write_csv_sheet("9. New Tables (proposed)", "proposed_tables.csv",
                "Phase 9 - New tables we may need",
                "Each new table is challenged before proposing. We only add a table if data truly cannot live elsewhere.")

write_csv_sheet("10. Relationships", "relationship_mapping.csv",
                "Phase 10 - How the tables link together",
                "Parent-to-child links and why they exist. 1:1 = one each, 1:M = one-to-many, M:M = many-to-many.")

# ==========================================================================
# Sheet: Big Risks
# ==========================================================================
risks = [
    ["Last year's revenue", "We multiply this year's number by a fixed ratio to guess last year.", "Growth % is wrong, so targets look easier/harder than reality.", "A real sales/ERP system with last completed year actuals.", "HIGH"],
    ["Last year's revenue by product (5 splits)", "We split last year using fixed ratios (18/32/15/20/15%).", "Per-product growth is fiction.", "Category-level actual sales.", "HIGH"],
    ["Current average order value (AOV)", "Hard-coded as 1,45,000 for EVERY employee.", "AOV growth score is meaningless.", "Sales actuals: revenue / number of orders.", "HIGH"],
    ["Current revenue per school", "Hard-coded as 2,40,000 for EVERY employee.", "Revenue-per-school growth is meaningless.", "Sales actuals / active school count.", "HIGH"],
    ["School counts (total/active/user/non-user)", "Made-up demo numbers (320/110/78/242).", "School growth and 'cost per school' are wrong.", "A school master list.", "HIGH"],
    ["Dashboard achievement % and YTD actual", "A random demo formula based on the employee code.", "The whole 'how am I doing' view is fake.", "A monthly actuals feed (year-to-date).", "HIGH"],
    ["Collection % per region", "Hard-coded map (North 85, West 88, South 90, East 86).", "Cash collection plan may be wrong per region.", "A Finance-owned region policy table.", "MEDIUM"],
    ["Validation thresholds", "60% growth = aggressive, cost > 25% = warning, etc.", "We may warn (or not warn) at the wrong levels.", "Business-confirmed rule values + an owner.", "MEDIUM"],
    ["Premium school strategy field", "It exists in the data model but is NOT shown on any screen.", "Either a missing screen or a dead field.", "Decision: show it or delete it.", "LOW"],
    ["Collection plan storage", "Step 7 is only in code; nothing is saved to a table.", "We cannot report on or lock the collection plan.", "Decision: add a collection table or always recompute.", "MEDIUM"],
]
ws = write_plain_sheet(
    "Big Risks",
    "The fake / unconfirmed values (read this first)",
    "These are the values that are currently made-up or hard-coded. Each must be resolved before go-live.",
    ["Value", "What we do today (the problem)", "Why it matters", "What it should come from", "Severity"],
    risks,
)
sev_fill = {"HIGH": STATUS_FILL["NEEDS_MAPPING"], "MEDIUM": STATUS_FILL["NEEDS_INPUT"], "LOW": STATUS_FILL["N/A"]}
for row in range(5, ws.max_row + 1):
    v = ws.cell(row=row, column=5).value
    if v in sev_fill:
        ws.cell(row=row, column=5).fill = sev_fill[v]

# ==========================================================================
# Sheet: Open Questions
# ==========================================================================
questions = [
    [1, "Last year revenue", "Which system/table holds last completed year revenue per employee, and which months count as 'last year'?", "revenue_targets.last_year_revenue + Revenue Growth %"],
    [2, "Last year by product", "Do we have real category-level sales for last year, or must we keep using a ratio?", "5 *_revenue_ly fields"],
    [3, "Current AOV", "Exact formula and time window for average order value (revenue / orders over how long)?", "current_aov + AOV Growth %"],
    [4, "Current revenue per school", "Top and bottom of the formula, and which schools count (active? user only?)?", "current_revenue_per_school"],
    [5, "School counts", "Is there a school master list? What defines total / active / user / non-user?", "Universe counts + category current counts"],
    [6, "Employee current revenue & target", "Source system for each employee's prior actual revenue AND their target?", "users.current_revenue, users.current_target, Home achievement %"],
    [7, "YTD actual & achievement %", "Which actuals feed, and at what level (total vs by product)?", "Dashboard YTD achievement"],
    [8, "Collection %", "Confirmed % per region + owner (Finance?). Is it % of revenue or of orders?", "collection_percent"],
    [9, "Collection phasing", "Are Dec/Mar/Apr/Jun cumulative shares (40/70/85/100%) the correct, owned schedule?", "4 collection milestones"],
    [10, "Store collection?", "Should Step 7 be saved (new table) or always recalculated on the fly?", "collection_planning table decision"],
    [11, "Revenue per school score", "Denominator: target schools or active schools? (code mixes both)", "Revenue/School KPI"],
    [12, "Validation thresholds", "Confirm 60% growth, x50 AOV ceiling, 25% cost limit - values + owner?", "Review flags"],
    [13, "Premium school strategy", "Keep and show it in the UI, or delete the field?", "universe_planning.premium_school_strategy"],
    [14, "Monthly phasing", "Adopt a monthly target split table, or accept a flat 1/12 split?", "revenue_phasing table decision"],
    [15, "Actuals ingestion", "Source, frequency, and level of detail for the monthly actuals feed?", "actuals_tracking feed"],
    [16, "Hiring approval", "Does hiring need its own approval chain, or is a single status enough?", "Hiring workflow"],
    [17, "Fiscal calendar", "Introduce a fiscal-calendar table now, or hard-code FY26-27?", "all 'current/last' logic"],
    [18, "Business owners", "Name an owner for each master/feed: school master, actuals/ERP, region policy, targets.", "governance"],
]
write_plain_sheet(
    "Open Questions",
    "Phase 12 - 18 questions we need YOU to answer",
    "Answering these unlocks the red (NEEDS_MAPPING) values. Each question maps to a specific field or score.",
    ["#", "Topic", "Question (please answer)", "Unlocks / maps to"],
    questions,
)

# ==========================================================================
# order + save
# ==========================================================================
order = ["Start Here", "1. Screens", "2. All Fields", "3. Where Values Come From",
         "4. KPIs", "5. Calculations", "6. Historical Data Needed", "7. Pre-fill Logic",
         "8. Existing Tables", "9. New Tables (proposed)", "10. Relationships",
         "Big Risks", "Open Questions"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", [s.title for s in wb._sheets])
